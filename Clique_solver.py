from multiprocessing.sharedctypes import SynchronizedBase
import symtable
from xmlrpc.client import boolean
from pysmt.shortcuts import Symbol, And, Not, Or, is_sat
import openpyxl
import numpy as np
import time

def Clique_solver(Query):
    # number of vertices in the graph
    n = 0
    matrix = []
    start_time = time.time()
    wookbook = openpyxl.load_workbook(Query[0])
    worksheet = wookbook.active

    # get the number of vertices
    n = worksheet.max_row
    # initlize a n*n matrix with 0s
    matrix = np.zeros((n,n))
    row_m = 0
    col_m = 0
    # reading data into the matrix
    for i in range(1, worksheet.max_row+1):
        for j in range(0, worksheet.max_column):
            matrix[row_m][col_m] = worksheet[i][j].value
            col_m+=1
        row_m += 1
        col_m = 0

    c = Query[1]
    if n < c:
        end_time = time.time()
        time_elapsed = (end_time - start_time)
        return (False, time_elapsed)

    if c == 0 or c == 1:
        end_time = time.time()
        time_elapsed = (end_time - start_time)
        return (True, time_elapsed)

    # building boolean formyla
    boolean_formula = Symbol('DummyVar')
    symbol_table = []
    # create 1A, 2A, 3A... 1B, 2B, 3B......
    for i in range(n):
        symbol_same_vertex = []
        for j in range(c):
            symbol_same_vertex.append(Symbol(str(j)+'_'+str(i)))
        symbol_table.append(symbol_same_vertex)

    # build boolean formula according to first constraint
    for j in range(c):
        formula_of_ors = ''
        for i in range(n-1):
            if formula_of_ors == '':
                formula_of_ors = Or(symbol_table[i][j], symbol_table[i+1][j])
            else:
                formula_of_ors = Or(formula_of_ors, symbol_table[i+1][j])
        boolean_formula = And(boolean_formula, formula_of_ors)
        #print(formula_of_ors)
        formula_of_ors = '' # reset
    #print()
    # build boolean formula according to second constraint
    for i in range(n):
        for j in range(c-1):
            for k in range(j+1, c):
                boolean_formula = And(boolean_formula, Or(Not(symbol_table[i][j]), Not(symbol_table[i][k])))
                #print(Or(Not(symbol_table[i][j]), Not(symbol_table[i][k])))
    #print()

    # build boolean formula according to third constraint
    for i in range(n):
        for j in range(n):
            if j>i and matrix[i][j] == 0: # (i, j) is a non-edge
                symbols_i = []
                symbols_j = []
                for k in range(c):
                    symbols_i.append(symbol_table[i][k])
                    symbols_j.append(symbol_table[j][k])
                for index1, k in enumerate(symbols_i):
                    for index2, m in enumerate(symbols_j):
                        if index1 != index2:
                            boolean_formula = And(boolean_formula, Or(Not(k), Not(m)))
                            #print(Or(Not(k), Not(m)))
                #print('-----')
    result = is_sat(boolean_formula)

    end_time = time.time()
    time_elapsed = (end_time - start_time)
    return (result, time_elapsed)